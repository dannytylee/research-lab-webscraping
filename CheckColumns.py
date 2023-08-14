import os
import pandas as pd
from openpyxl import load_workbook

# Set the directory where your Excel files are located
directory = r'C:\Users\danny\OneDrive\Files\GitHub\research-lab-webscraping\Data'

# Define expected column names
expected_columns = ['FIPS', 'Name', '2013 Rural-urban Continuum Code*', '1970', '1980', '1990', '2000', '2008-2012', '2017-2021']

# Initialize a list to store file names with incorrect columns
files_with_incorrect_columns = []

# Iterate through all Excel files in the directory
for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(directory, filename)

        # Load the workbook using openpyxl
        wb = load_workbook(file_path, data_only=True)  # data_only=True to get cell values, not formulas
        ws = wb.active

        # Read the third row which contains column names
        header_row = ws[3]
        actual_columns = [cell.value for cell in header_row]

        if actual_columns != expected_columns:
            files_with_incorrect_columns.append(filename)
            print(f"Warning: {filename} has incorrect columns - {actual_columns}")

# Print the files with incorrect columns
if files_with_incorrect_columns:
    print("\nFiles with incorrect columns:")
    for file in files_with_incorrect_columns:
        print(file)
else:
    print("All files have correct columns.")